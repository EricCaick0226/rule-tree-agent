from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_path_level(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"^[A-Z][、.]?", "", text)
    if not re.match(r"^\d+(周岁|岁)", text):
        text = re.sub(r"^\d+[、.]?", "", text)
    return text.strip()


def normalize_grade(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("一般数据", "")
    text = re.sub(r"\s+", "", text)
    return text


def _key(row: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        normalized
        for level in row.get("path_levels", [])
        if (normalized := normalize_path_level(level))
    )


def compare_rows(
    generated: list[dict[str, Any]],
    reference: list[dict[str, Any]],
) -> dict[str, Any]:
    generated_by_key = {_key(row): row for row in generated if _key(row)}
    reference_by_key = {_key(row): row for row in reference if _key(row)}
    matched_keys = [key for key in generated_by_key if key in reference_by_key]

    grade_matches = 0
    grade_mismatches = 0
    grade_mismatch_examples: list[dict[str, Any]] = []
    for key in matched_keys:
        generated_grade = normalize_grade(generated_by_key[key].get("recommended_grade"))
        reference_grade = normalize_grade(reference_by_key[key].get("recommended_grade"))
        if generated_grade == reference_grade:
            grade_matches += 1
            continue
        grade_mismatches += 1
        if len(grade_mismatch_examples) < 20:
            grade_mismatch_examples.append(
                {
                    "path_levels": list(key),
                    "generated_grade": generated_grade,
                    "reference_grade": reference_grade,
                }
            )

    missing_keys = [key for key in reference_by_key if key not in generated_by_key]
    extra_keys = [key for key in generated_by_key if key not in reference_by_key]
    return {
        "generated_rows": len(generated_by_key),
        "reference_rows": len(reference_by_key),
        "matched_paths": len(matched_keys),
        "coverage_pct": round(len(matched_keys) / max(1, len(reference_by_key)) * 100, 2),
        "missing_paths": len(missing_keys),
        "extra_paths": len(extra_keys),
        "grade_matches": grade_matches,
        "grade_mismatches": grade_mismatches,
        "grade_accuracy_on_matched_pct": round(
            grade_matches / max(1, len(matched_keys)) * 100,
            2,
        ),
        "missing_path_examples": [list(key) for key in missing_keys[:20]],
        "extra_path_examples": [list(key) for key in extra_keys[:20]],
        "grade_mismatch_examples": grade_mismatch_examples,
    }


def load_reference_rows(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["分类"]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        path_levels = [normalize_path_level(value) for value in values[:5]]
        path_levels = [level for level in path_levels if level]
        if not path_levels:
            continue
        rows.append(
            {
                "path_levels": path_levels,
                "recommended_grade": normalize_grade(values[5]),
                "description": "" if values[6] is None else str(values[6]).strip(),
            }
        )
    workbook.close()
    return rows


def load_generated_rows(rule_table_path: Path) -> list[dict[str, Any]]:
    data = json.loads(rule_table_path.read_text(encoding="utf-8"))
    return list(data.get("classification_rows") or [])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compare generated rule_table.json against reference Excel.",
    )
    parser.add_argument("--generated", required=True, help="Path to generated rule_table.json")
    parser.add_argument("--reference", required=True, help="Path to reference xlsx")
    args = parser.parse_args()

    result = compare_rows(
        load_generated_rows(Path(args.generated)),
        load_reference_rows(Path(args.reference)),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
