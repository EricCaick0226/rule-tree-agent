from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


INSUFFICIENT_DESCRIPTION = "证据不足，无法从当前文档确定"
DESCRIPTION_SOURCE = "classification_standard_excel"
SOURCE_CONFIDENCE = "curated_answer"
STRONGER_DESCRIPTION_SOURCES = {
    "local_standard_quote",
    "national_standard_quote",
    "standard_quote",
    "quoted_standard",
    "quoted_local_regulation",
}


def normalize_path_level(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"^[A-Z][、.]", "", text)
    if not re.match(r"^\d+(周岁|岁)", text):
        text = re.sub(r"^\d+[、.]", "", text)
    return text.strip()


def _path_key(path_levels: list[Any]) -> tuple[str, ...]:
    return tuple(
        normalized
        for level in path_levels
        if (normalized := normalize_path_level(level))
    )


def _stable_excel_row_id(path_key: tuple[str, ...]) -> str:
    digest = hashlib.sha1("/".join(path_key).encode("utf-8")).hexdigest()[:12]
    return f"row_excel_{digest}"


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def load_excel_rows(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["分类"]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        path_levels = [normalize_path_level(value) for value in values[:5]]
        path_levels = [level for level in path_levels if level]
        description = "" if values[6] is None else str(values[6]).strip()
        if not path_levels:
            continue
        rows.append(
            {
                "row_number": row_number,
                "path_levels": path_levels,
                "path_key": tuple(path_levels),
                "recommended_grade": "" if values[5] is None else str(values[5]).strip(),
                "description": description,
            }
        )
    workbook.close()
    return rows


def _has_stronger_description(row: dict[str, Any]) -> bool:
    source = str(row.get("description_source") or "").strip()
    quote = str(row.get("description_evidence_quote") or "").strip()
    description = str(row.get("description") or "").strip()
    if not description or description == INSUFFICIENT_DESCRIPTION:
        return False
    if source in STRONGER_DESCRIPTION_SOURCES:
        return True
    return bool(quote and source and source != DESCRIPTION_SOURCE)


def _already_current_from_excel(row: dict[str, Any], description: str) -> bool:
    return (
        str(row.get("description") or "").strip() == description
        and str(row.get("description_source") or "").strip() == DESCRIPTION_SOURCE
        and str(row.get("source_confidence") or "").strip() == SOURCE_CONFIDENCE
        and "description_evidence_quote" not in row
    )


def import_descriptions(
    rule_table_path: Path,
    excel_path: Path,
    report_path: Path | None = None,
) -> dict[str, Any]:
    data = _load_json(rule_table_path)
    rows = data.setdefault("classification_rows", [])
    excel_rows = load_excel_rows(excel_path)
    excel_path_counts = Counter(row["path_key"] for row in excel_rows)
    excel_by_path = {
        row["path_key"]: row
        for row in excel_rows
        if excel_path_counts[row["path_key"]] == 1
    }
    ambiguous_paths = [
        list(path_key)
        for path_key, count in sorted(excel_path_counts.items())
        if count > 1
    ]
    ambiguous_excel_rows = sum(
        count
        for count in excel_path_counts.values()
        if count > 1
    )

    exact_path_matches = 0
    descriptions_imported = 0
    skipped_already_current = 0
    skipped_existing_stronger_source = 0
    matched_excel_paths: set[tuple[str, ...]] = set()
    imported_rows: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    existing_path_keys: set[tuple[str, ...]] = set()

    for row in rows:
        if not isinstance(row, dict):
            continue
        path_key = _path_key(list(row.get("path_levels") or []))
        if not path_key:
            continue
        existing_path_keys.add(path_key)
        excel_row = excel_by_path.get(path_key)
        if excel_row is None:
            continue
        exact_path_matches += 1
        matched_excel_paths.add(path_key)
        description = str(excel_row.get("description") or "").strip()
        if not description:
            skipped_rows.append({"reason": "empty_excel_description", "path_levels": list(path_key)})
            continue
        if _already_current_from_excel(row, description):
            skipped_already_current += 1
            skipped_rows.append({"reason": "already_current", "path_levels": list(path_key)})
            continue
        if _has_stronger_description(row):
            skipped_existing_stronger_source += 1
            skipped_rows.append({"reason": "existing_stronger_source", "path_levels": list(path_key)})
            continue

        row["description"] = description
        row["description_source"] = DESCRIPTION_SOURCE
        row["source_confidence"] = SOURCE_CONFIDENCE
        row.pop("description_evidence_quote", None)
        descriptions_imported += 1
        imported_rows.append(
            {
                "row_id": row.get("row_id"),
                "path_levels": list(path_key),
                "excel_row_number": excel_row.get("row_number"),
            }
        )

    new_rows_appended = 0
    appended_rows: list[dict[str, Any]] = []
    for path_key, excel_row in sorted(excel_by_path.items()):
        if path_key in existing_path_keys:
            continue
        description = str(excel_row.get("description") or "").strip()
        if not description:
            skipped_rows.append({"reason": "empty_excel_description", "path_levels": list(path_key)})
            continue
        new_row = {
            "row_id": _stable_excel_row_id(path_key),
            "path_levels": list(path_key),
            "recommended_grade": str(excel_row.get("recommended_grade") or "").strip(),
            "description": description,
            "description_source": DESCRIPTION_SOURCE,
            "source_confidence": SOURCE_CONFIDENCE,
            "row_source": DESCRIPTION_SOURCE,
            "curation_status": "classification_standard_excel_import",
        }
        rows.append(new_row)
        existing_path_keys.add(path_key)
        new_rows_appended += 1
        appended_rows.append(
            {
                "row_id": new_row["row_id"],
                "path_levels": list(path_key),
                "excel_row_number": excel_row.get("row_number"),
            }
        )

    skipped_no_exact_reference_path = sum(
        1
        for excel_row in excel_rows
        if excel_row["path_key"] not in matched_excel_paths
        and excel_row["path_key"] not in existing_path_keys
        and excel_path_counts[excel_row["path_key"]] == 1
    )

    _write_json(rule_table_path, data)
    summary = {
        "rule_table": str(rule_table_path),
        "excel": str(excel_path),
        "excel_rows": len(excel_rows),
        "exact_path_matches": exact_path_matches,
        "descriptions_imported": descriptions_imported,
        "new_rows_appended": new_rows_appended,
        "skipped_already_current": skipped_already_current,
        "skipped_existing_stronger_source": skipped_existing_stronger_source,
        "skipped_no_exact_reference_path": skipped_no_exact_reference_path,
        "ambiguous_excel_paths": len(ambiguous_paths),
        "ambiguous_excel_rows": ambiguous_excel_rows,
        "skipped_ambiguous_excel_rows": ambiguous_excel_rows,
        "ambiguous_path_examples": ambiguous_paths[:20],
        "imported_rows": imported_rows,
        "appended_rows": appended_rows,
        "skipped_rows": skipped_rows[:50],
    }
    if report_path is not None:
        _write_json(report_path, summary)
    return summary


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import curated descriptions from 分类分级标准.xlsx into exact reference rows.",
    )
    parser.add_argument(
        "--rule-table",
        default="reference_library/wst787_2021/rule_table.json",
        help="Reference rule_table.json to update.",
    )
    parser.add_argument(
        "--excel",
        default="data/input_docs/分类分级标准.xlsx",
        help="Classification standard Excel workbook.",
    )
    parser.add_argument(
        "--report",
        default="reference_library/wst787_2021/import_reports/classification_standard_description_import.json",
        help="Path to write merge report JSON.",
    )
    args = parser.parse_args(argv)
    summary = import_descriptions(Path(args.rule_table), Path(args.excel), Path(args.report))
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
